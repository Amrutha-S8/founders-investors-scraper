"""
Indian Startup Investment Scraper & Analyzer
=============================================
Reads a raw CSV of Indian startup funding data,
cleans it, analyzes it, and exports a formatted
multi-sheet Excel workbook.

Data: 3,043 startup investment deals (2015-2020)
Output: startups_cleaned.xlsx

Run:
    pip install -r requirements.txt
    python scraper.py
"""

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


# ──────────────────────────────────────────────
#  STEP 1: Load & Clean Data
# ──────────────────────────────────────────────

def load_and_clean(filepath="database.csv"):
    df = pd.read_csv(filepath)

    df = df.rename(columns={
        "Startup Name":      "Company Name",
        "Investors Name":    "Investor(s)",
        "Date dd/mm/yyyy":   "Date",
        "Industry Vertical": "Industry",
        "SubVertical":       "Sub-Vertical",
        "City  Location":    "City",
        "InvestmentnType":   "Investment Type",
        "Amount in USD":     "Amount (USD)",
    })

    df = df.drop(columns=["Sr No", "Remarks"], errors="ignore")

    # Fix Amount: remove commas and convert to float
    df["Amount (USD)"] = (
        df["Amount (USD)"]
        .astype(str)
        .str.replace(",", "", regex=False)
        .str.strip()
    )
    df["Amount (USD)"] = pd.to_numeric(df["Amount (USD)"], errors="coerce")

    # Normalize Investment Type labels
    replacements = {
        "Seed/ Angel Funding":  "Seed/Angel Funding",
        "Seed / Angel Funding": "Seed/Angel Funding",
        "Seed\nFunding":        "Seed Funding",
        "Seed\\nFunding":       "Seed Funding",
    }
    df["Investment Type"] = (
        df["Investment Type"].astype(str).str.strip().replace(replacements)
    )

    # Drop rows where company name is a URL (dirty data)
    df = df[~df["Company Name"].astype(str).str.startswith("http")]

    # Strip whitespace from key string columns
    for col in ["Company Name", "Investor(s)", "City", "Industry"]:
        df[col] = df[col].astype(str).str.strip()

    df = df.reset_index(drop=True)
    print(f"  Loaded {len(df):,} clean records")
    return df


# ──────────────────────────────────────────────
#  STEP 2: Excel Styling Helpers
# ──────────────────────────────────────────────

def thin_border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)


def header_cell(cell, text, bg="1F3864", fg="FFFFFF", sz=11, bold=True):
    cell.value = text
    cell.font  = Font(name="Arial", bold=bold, color=fg, size=sz)
    cell.fill  = PatternFill("solid", start_color=bg, end_color=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = thin_border()


def data_cell(cell, value, align="left", bold=False, bg=None, fmt=None):
    cell.value = value
    cell.font  = Font(name="Arial", size=10, bold=bold)
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
    cell.border = thin_border()
    if bg:
        cell.fill = PatternFill("solid", start_color=bg, end_color=bg)
    if fmt:
        cell.number_format = fmt


# ──────────────────────────────────────────────
#  STEP 3: Build Excel Workbook
# ──────────────────────────────────────────────

def build_excel(df, output="startups_cleaned.xlsx"):
    wb    = openpyxl.Workbook()
    total = len(df)

    # ── Sheet 1: Dashboard ──────────────────────
    ws1 = wb.active
    ws1.title = "Dashboard"
    ws1.sheet_view.showGridLines = False

    ws1.merge_cells("A1:H1")
    header_cell(
        ws1["A1"],
        f"Indian Startup Investment Dashboard  |  {total:,} Deals  |  $38B+ Invested",
        bg="1F3864", sz=14
    )
    ws1.row_dimensions[1].height = 38

    # KPI cards
    ws1.row_dimensions[3].height = 55
    kpis = [
        ("A3:B3", f"{total:,}",                              "Total Deals",      "E8F0FE", "1F3864"),
        ("C3:D3", f"${df['Amount (USD)'].sum()/1e9:.1f}B",   "Total Invested",   "E8F5E9", "1B5E20"),
        ("E3:F3", f"${df['Amount (USD)'].mean()/1e6:.1f}M",  "Avg Deal Size",    "F3E5F5", "4A148C"),
        ("G3:H3", f"{df['City'].nunique()} Cities",           "Cities Covered",   "FFEBEE", "B71C1C"),
    ]
    for rng, val, lbl, bg, col in kpis:
        ws1.merge_cells(rng)
        c = ws1[rng.split(":")[0]]
        c.value = val
        c.font  = Font(name="Arial", bold=True, size=20, color=col)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.fill  = PatternFill("solid", start_color=bg, end_color=bg)
        c.border = thin_border()

    ws1.row_dimensions[4].height = 18
    for rng, lbl, col in [
        ("A4:B4","Total Deals","1F3864"),("C4:D4","Total Invested","1B5E20"),
        ("E4:F4","Avg Deal Size","4A148C"),("G4:H4","Cities Covered","B71C1C"),
    ]:
        ws1.merge_cells(rng)
        c = ws1[rng.split(":")[0]]
        c.value = lbl
        c.font  = Font(name="Arial", size=9, bold=True, color=col)
        c.alignment = Alignment(horizontal="center")

    # Top Cities table
    ws1.merge_cells("A6:D6")
    header_cell(ws1["A6"], "Top 10 Cities by Deal Count", bg="2E4057")
    ws1.row_dimensions[6].height = 22
    for j, h in enumerate(["City","Deals","Total Invested (USD)","% Share"], 1):
        header_cell(ws1.cell(7, j), h, bg="3A6186", sz=10)
    ws1.row_dimensions[7].height = 20

    city_grp = (
        df.groupby("City")
        .agg(deals=("Company Name","count"), total=("Amount (USD)","sum"))
        .sort_values("deals", ascending=False)
        .head(10)
    )
    for i, (city, row) in enumerate(city_grp.iterrows()):
        r  = i + 8
        bg = "F0F4FF" if i % 2 == 0 else "FFFFFF"
        data_cell(ws1.cell(r,1), city, bold=True, bg=bg)
        data_cell(ws1.cell(r,2), int(row["deals"]), "center", bg=bg)
        data_cell(ws1.cell(r,3), float(row["total"]) if pd.notna(row["total"]) else 0, "right", bg=bg, fmt="$#,##0")
        data_cell(ws1.cell(r,4), f"{row['deals']/total*100:.1f}%", "center", bg=bg)

    # Top Industries table
    ws1.merge_cells("F6:H6")
    header_cell(ws1["F6"], "Top 10 Industries", bg="2E4057")
    for j, h in enumerate(["Industry","Deals","% Share"], 1):
        header_cell(ws1.cell(7, j+5), h, bg="3A6186", sz=10)

    ind_grp = df.groupby("Industry").size().sort_values(ascending=False).head(10)
    for i, (ind, cnt) in enumerate(ind_grp.items()):
        r  = i + 8
        bg = "FFF8E1" if i % 2 == 0 else "FFFFFF"
        data_cell(ws1.cell(r,6), ind, bold=True, bg=bg)
        data_cell(ws1.cell(r,7), int(cnt), "center", bg=bg)
        data_cell(ws1.cell(r,8), f"{cnt/total*100:.1f}%", "center", bg=bg)

    for col, w in [("A",18),("B",9),("C",18),("D",10),("E",3),("F",22),("G",9),("H",10)]:
        ws1.column_dimensions[col].width = w

    # ── Sheet 2: All Deals ──────────────────────
    ws2 = wb.create_sheet("All Deals")
    ws2.sheet_view.showGridLines = False
    ws2.freeze_panes = "A3"
    ws2.merge_cells("A1:H1")
    header_cell(ws2["A1"], f"Complete Indian Startup Investment Database  |  {total:,} Records", bg="1F3864", sz=13)
    ws2.row_dimensions[1].height = 30

    hdrs2 = ["Date","Company Name","Industry","City","Investor(s)","Investment Type","Amount (USD)","Sub-Vertical"]
    for j, h in enumerate(hdrs2, 1):
        header_cell(ws2.cell(2, j), h, bg="2E4057")
    ws2.row_dimensions[2].height = 22

    cols2 = ["Date","Company Name","Industry","City","Investor(s)","Investment Type","Amount (USD)","Sub-Vertical"]
    for i, row_data in df[cols2].iterrows():
        r  = i + 3
        bg = "F0F4FF" if i % 2 == 0 else "FFFFFF"
        data_cell(ws2.cell(r,1), str(row_data["Date"]), bg=bg)
        data_cell(ws2.cell(r,2), str(row_data["Company Name"]), bold=True, bg=bg)
        data_cell(ws2.cell(r,3), str(row_data["Industry"]), bg=bg)
        data_cell(ws2.cell(r,4), str(row_data["City"]), bg=bg)
        data_cell(ws2.cell(r,5), str(row_data["Investor(s)"]), bg=bg)
        data_cell(ws2.cell(r,6), str(row_data["Investment Type"]), bg=bg)
        amt = row_data["Amount (USD)"]
        data_cell(ws2.cell(r,7), float(amt) if pd.notna(amt) else "", "right", bg=bg, fmt="$#,##0")
        sv  = row_data["Sub-Vertical"]
        data_cell(ws2.cell(r,8), str(sv) if pd.notna(sv) else "", bg=bg)

    ws2.auto_filter.ref = f"A2:H{total+2}"
    for col, w in [("A",13),("B",24),("C",18),("D",14),("E",30),("F",18),("G",15),("H",22)]:
        ws2.column_dimensions[col].width = w

    # ── Sheet 3: Investor Analysis ──────────────
    ws3 = wb.create_sheet("Investor Analysis")
    ws3.sheet_view.showGridLines = False
    ws3.merge_cells("A1:F1")
    header_cell(ws3["A1"], "Top 50 Investors by Deal Count  |  Indian Startup Ecosystem", bg="1B5E20", sz=13)
    ws3.row_dimensions[1].height = 30

    for j, h in enumerate(["Investor Name","Total Deals","Total Invested (USD)","Avg Deal (USD)","Top City","Top Industry"], 1):
        header_cell(ws3.cell(2, j), h, bg="2E7D32")
    ws3.row_dimensions[2].height = 22

    inv_grp = (
        df.groupby("Investor(s)")
        .agg(deals=("Company Name","count"), total=("Amount (USD)","sum"), avg=("Amount (USD)","mean"))
        .sort_values("deals", ascending=False)
        .head(50)
    )
    for i, (inv, row) in enumerate(inv_grp.iterrows()):
        r      = i + 3
        bg     = "F1F8E9" if i % 2 == 0 else "FFFFFF"
        inv_df = df[df["Investor(s)"] == inv]
        vc     = inv_df["City"].value_counts()
        vi     = inv_df["Industry"].value_counts()
        top_city = vc.index[0] if len(vc) > 0 else ""
        top_ind  = vi.index[0] if len(vi) > 0 else ""
        data_cell(ws3.cell(r,1), str(inv), bold=True, bg=bg)
        data_cell(ws3.cell(r,2), int(row["deals"]), "center", bg=bg)
        data_cell(ws3.cell(r,3), float(row["total"]) if pd.notna(row["total"]) else 0, "right", bg=bg, fmt="$#,##0")
        data_cell(ws3.cell(r,4), float(row["avg"])   if pd.notna(row["avg"])   else 0, "right", bg=bg, fmt="$#,##0")
        data_cell(ws3.cell(r,5), str(top_city), bg=bg)
        data_cell(ws3.cell(r,6), str(top_ind),  bg=bg)

    for col, w in [("A",32),("B",13),("C",22),("D",18),("E",16),("F",20)]:
        ws3.column_dimensions[col].width = w

    # ── Sheet 4: Investment Stages ──────────────
    ws4 = wb.create_sheet("Investment Stages")
    ws4.sheet_view.showGridLines = False
    ws4.merge_cells("A1:E1")
    header_cell(ws4["A1"], "Investment Stage & Type Breakdown", bg="4A148C", sz=13)
    ws4.row_dimensions[1].height = 30

    for j, h in enumerate(["Investment Type","Deal Count","% of Total","Total Amount (USD)","Avg Amount (USD)"], 1):
        header_cell(ws4.cell(2, j), h, bg="6A1B9A")
    ws4.row_dimensions[2].height = 22

    stage_grp = (
        df.groupby("Investment Type")
        .agg(count=("Company Name","count"), total=("Amount (USD)","sum"), avg=("Amount (USD)","mean"))
        .sort_values("count", ascending=False)
    )
    for i, (stage, row) in enumerate(stage_grp.iterrows()):
        r  = i + 3
        bg = "F3E5F5" if i % 2 == 0 else "FFFFFF"
        data_cell(ws4.cell(r,1), str(stage), bold=True, bg=bg)
        data_cell(ws4.cell(r,2), int(row["count"]), "center", bg=bg)
        data_cell(ws4.cell(r,3), f"{row['count']/total*100:.1f}%", "center", bg=bg)
        data_cell(ws4.cell(r,4), float(row["total"]) if pd.notna(row["total"]) else 0, "right", bg=bg, fmt="$#,##0")
        data_cell(ws4.cell(r,5), float(row["avg"])   if pd.notna(row["avg"])   else 0, "right", bg=bg, fmt="$#,##0")

    for col, w in [("A",22),("B",13),("C",16),("D",22),("E",20)]:
        ws4.column_dimensions[col].width = w

    wb.save(output)
    print(f"  Saved → {output}")


# ──────────────────────────────────────────────
#  MAIN
# ──────────────────────────────────────────────

if __name__ == "__main__":
    print("\n Indian Startup Investment Scraper & Analyzer")
    print("=" * 48)

    print("\n[1/2] Loading and cleaning data...")
    df = load_and_clean("database.csv")

    print("\n[2/2] Building Excel workbook...")
    build_excel(df, "startups_cleaned.xlsx")

    print("\nDone!")
    print(f"  Records  : {len(df):,}")
    print(f"  Output   : startups_cleaned.xlsx  (4 sheets)")
    print(f"  Sheets   : Dashboard | All Deals | Investor Analysis | Investment Stages")
